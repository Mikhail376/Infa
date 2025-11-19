import pandas as pd
import polars as pl
from openpyxl import Workbook
from openpyxl.styles import Border, Side

# Чтение Excel через Pandas, замена NaN на ""
df_pd = pd.read_excel("Лабораторная 5.xlsm", header=None).fillna("").astype(str)

# Конвертация в Polars
df = pl.from_pandas(df_pd)

# Срез строк 4–15 (индексы 3–14)
df_sliced = df.slice(3, 12)

# Выбор колонок A–Y (0–24) без F (индекс 5)
selected_cols = [str(i) for i in range(25) if i != 5]
selected = df_sliced.select([pl.col(col) for col in selected_cols])

# Убираем Null в Polars
selected = selected.fill_null("")

# Создание Excel через OpenPyXL
wb = Workbook()
ws = wb.active

# Перенос значений
for r, row in enumerate(selected.rows(), start=1):
    for c, value in enumerate(row, start=1):
        ws.cell(row=r, column=c).value = value

# Определяем границы
thin = Side(border_style="thin", color="000000")
medium = Side(border_style="medium", color="000000")  # для толстых вертикальных линий

max_row = selected.height
max_col = selected.width

for r in range(1, max_row + 1):
    for c in range(1, max_col + 1):
        # Внешний прямоугольник
        left = thin if c == 1 else None
        right = thin if c == max_col else None
        top = thin if r == 1 else None
        bottom = thin if r == max_row else None

        # Толстые линии между A:B:C
        if c in [1, 2]:  # вертикальные линии между A|B и B|C
            right = medium

        # Толстые линии с обеих сторон блока D:E
        if c == 4:  # левая граница блока D
            left = medium
        if c == 5:  # правая граница блока E
            right = medium

        ws.cell(row=r, column=c).border = Border(left=left, right=right, top=top, bottom=bottom)

# Сохраняем Excel
wb.save("polars_rectangle_abc_de.xlsx")
print("Файл 'polars_rectangle_abc_de.xlsx' успешно создан с выделением столбцов A-C и D-E!")
